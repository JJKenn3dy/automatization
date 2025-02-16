import sys
import os
import getpass
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget,
    QVBoxLayout, QHBoxLayout, QSizePolicy,
    QLabel, QPushButton, QStackedWidget,
    QTableWidget, QTableWidgetItem, QSystemTrayIcon, QStyle, QDialog, QDialogButtonBox,
    QHeaderView
)
from PyQt6.QtSvgWidgets import QSvgWidget
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont, QFontDatabase, QIcon
from openpyxl import Workbook, load_workbook  # Excel
from openpyxl.styles import PatternFill, Font  # Стилизация ячеек
from documents import Type1


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DOM RF")
        self.resize(1200, 700)

        # 1) Создаем QStackedWidget
        self.stacked_widget = QStackedWidget()
        self.setCentralWidget(self.stacked_widget)

        # 2) Создаем и добавляем страницы
        self.page1 = self.create_page1()
        self.page2 = self.create_page2()
        self.page3 = self.create_page3()
        self.page4 = self.create_page4()
        self.page5 = self.create_page5()

        self.stacked_widget.addWidget(self.page1)  # Индекс 0
        self.stacked_widget.addWidget(self.page2)  # Индекс 1
        self.stacked_widget.addWidget(self.page3)  # Индекс 2
        self.stacked_widget.addWidget(self.page4)  # Индекс 3
        self.stacked_widget.addWidget(self.page5)  # Индекс 4

        # По умолчанию показываем первую страницу
        self.stacked_widget.setCurrentIndex(0)

    def create_page1(self) -> QWidget:
        """Создаём первую страницу: текст, несколько кнопок, SVG справа."""
        page = QWidget()
        main_layout = QHBoxLayout(page)
        main_layout.setContentsMargins(50, 40, 50, 40)

        # Левая колонка (текст + несколько кнопок)
        left_layout = QVBoxLayout()

        # Текст
        text_label = QLabel(f"Добро пожаловать {getpass.getuser()}!")
        text_label.setWordWrap(True)
        text_label.setStyleSheet("font-size: 30px; color: #62961e;")
        # Растягиваем текст по вертикали при необходимости
        text_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        left_layout.addWidget(text_label, alignment=Qt.AlignmentFlag.AlignTop)

        # Текст
        text_label = QLabel("Програма для автоматизации\nпроцессов ИБ")
        text_label.setWordWrap(True)
        text_label.setStyleSheet("font-size: 25px; color: #76787A;")
        # Растягиваем текст по вертикали при необходимости
        text_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        left_layout.addWidget(text_label, alignment=Qt.AlignmentFlag.AlignTop)

        # Добавим несколько кнопок
        btn_to_second = QPushButton(f"Лицензии")
        btn_to_second.clicked.connect(self.on_toggle)
        btn_to_second.setFixedSize(150, 50)
        btn_to_second.setStyleSheet('font-size: 15px; color: rgb(98, 150, 30);')
        btn_to_second.clicked.connect(self.go_to_second_page)
        left_layout.addWidget(btn_to_second, alignment=Qt.AlignmentFlag.AlignLeft)
        btn_to_third = QPushButton(f"СКЗИ")
        btn_to_third.clicked.connect(self.on_toggle)
        btn_to_third.setFixedSize(150, 50)
        btn_to_third.clicked.connect(self.go_to_third_page)
        btn_to_third.setStyleSheet('font-size: 15px; color: rgb(98, 150, 30);')
        left_layout.addWidget(btn_to_third, alignment=Qt.AlignmentFlag.AlignLeft)
        btn_to_fourth = QPushButton(f"Страница 3")
        btn_to_fourth.clicked.connect(self.on_toggle)
        btn_to_fourth.setFixedSize(150, 50)
        btn_to_fourth.setStyleSheet('font-size: 15px; color: rgb(98, 150, 30);')
        btn_to_fourth.clicked.connect(self.go_to_fourth_page)
        left_layout.addWidget(btn_to_fourth, alignment=Qt.AlignmentFlag.AlignLeft)
        btn_to_five = QPushButton(f"Страница 4")
        btn_to_five.clicked.connect(self.on_toggle)
        btn_to_five.setFixedSize(150, 50)
        btn_to_five.setStyleSheet('font-size: 15px; color: rgb(98, 150, 30);')
        btn_to_five.clicked.connect(self.go_to_five_page)
        left_layout.addWidget(btn_to_five, alignment=Qt.AlignmentFlag.AlignLeft)
        main_layout.addLayout(left_layout)

        # Справа – SVG
        self.svg_widget = QSvgWidget("logo.svg")
        self.svg_widget.setFixedSize(300, 300)
        self.svg_widget.setAutoFillBackground(True)
        # Обёртка для SVG
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(20, 20, 20, 20)
        container_layout.addWidget(self.svg_widget, alignment=Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(container, alignment=Qt.AlignmentFlag.AlignRight)

        return page

    def create_page2(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)

        if os.path.exists('Журналы2.xlsx'):
            wb = load_workbook('Журналы2.xlsx')
            ws = wb.active
            data = []

            # Считываем данные, принудительно задавая диапазон до 5 столбцов
            for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                data.append(list(row))

            if data:
                headers = data[0]
                # Если заголовков меньше 5, дополняем пустыми строками
                if len(headers) < 4:
                    headers += [""] * (4 - len(headers))
                rows = data[1:]

                table = QTableWidget()
                table.setColumnCount(4)  # Жестко задаем 5 столбцов
                table.setRowCount(len(rows))
                table.setHorizontalHeaderLabels(
                    [str(header) if header is not None else "" for header in headers]
                )

                for row_index, row in enumerate(rows):
                    # Если в строке меньше 5 ячеек, дополняем до 5 элементов
                    if len(row) < 4:
                        row += [None] * (4 - len(row))
                    for col_index, cell in enumerate(row):
                        item = QTableWidgetItem(str(cell) if cell is not None else "")
                        table.setItem(row_index, col_index, item)

                # Растягиваем столбцы, чтобы все 5 были видны равномерно
                table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
                table.setMinimumSize(1000, 500)
                layout.addWidget(table, alignment=Qt.AlignmentFlag.AlignCenter)
        else:
            QBtn = QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
            self.buttonBox = QDialogButtonBox(QBtn)
            self.buttonBox.accepted.connect(self.accept)
            self.buttonBox.rejected.connect(self.reject)
            layout.addWidget(self.buttonBox, alignment=Qt.AlignmentFlag.AlignLeft)

        btn_back = QPushButton("Назад")
        btn_back.clicked.connect(self.go_to_first_page)
        layout.addWidget(btn_back, alignment=Qt.AlignmentFlag.AlignLeft)
        return page


    def create_page3(self) -> QWidget:
        """Третья страница"""
        page = QWidget()
        layout = QVBoxLayout(page)

        label = QLabel("СКЗИ")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        btn_back = QPushButton("Назад на главную.")
        btn_back.clicked.connect(self.go_to_first_page)

        layout.addWidget(label)
        layout.addWidget(btn_back, alignment=Qt.AlignmentFlag.AlignCenter)

        return page

    def create_page4(self) -> QWidget:
        """Четвертая страница"""
        page = QWidget()
        layout = QVBoxLayout(page)

        label = QLabel("Это четвертая страница")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        btn_back = QPushButton("Назад на главную.")
        btn_back.clicked.connect(self.go_to_first_page)

        layout.addWidget(label)
        layout.addWidget(btn_back, alignment=Qt.AlignmentFlag.AlignCenter)

        return page

    def create_page5(self) -> QWidget:
        """Пятая страница"""
        page = QWidget()
        layout = QVBoxLayout(page)

        label = QLabel("Это пятая страница")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        btn_back = QPushButton("Назад на главную.")
        btn_back.clicked.connect(self.go_to_first_page)

        layout.addWidget(label)
        layout.addWidget(btn_back, alignment=Qt.AlignmentFlag.AlignCenter)

        return page

    def on_toggle(self):
        """Метод, который вызывается при нажатии на кнопки с первой страницы."""

    def go_to_five_page(self):
        """Переключиться на вторую страницу (индекс 4)."""
        self.stacked_widget.setCurrentIndex(4)

    def go_to_fourth_page(self):
        """Переключиться на вторую страницу (индекс 3)."""
        self.stacked_widget.setCurrentIndex(3)

    def go_to_third_page(self):
        """Переключиться на вторую страницу (индекс 2)."""
        self.stacked_widget.setCurrentIndex(2)

    def go_to_second_page(self):
        """Переключиться на вторую страницу (индекс 1)."""
        self.stacked_widget.setCurrentIndex(1)

    def go_to_first_page(self):
        """Вернуться на первую страницу (индекс 0)."""
        self.stacked_widget.setCurrentIndex(0)

    def accept(self):
        # Реализуйте нужное поведение, например, закрытие окна:
        self.close()

    def reject(self):
        # Реализуйте нужное поведение, например, закрытие окна:
        self.close()
def main():
    app = QApplication(sys.argv)



    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()
